import csv
import json
from typing import Any, Dict, Optional, cast
from django.contrib import admin
from django.core.exceptions import PermissionDenied
from django.http import Http404, HttpResponse
from django.urls import path, reverse
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill

from .models import WebsiteAudit, WebsiteAuditResult


def _format_dt(value) -> str:
    return value.isoformat() if value else ""


def _jsonify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True)
    return str(value)


@admin.register(WebsiteAudit)
class WebsiteAuditAdmin(admin.ModelAdmin):
    list_display = ("id", "url", "rendered", "status_code", "score", "fetched_at")
    list_filter = ("rendered", "status_code")
    search_fields = ("url",)
    actions = ["export_selected_csv", "export_selected_excel"]

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "<path:object_id>/export-csv/",
                self.admin_site.admin_view(self.export_csv),
                name="audits_websiteaudit_export_csv",
            ),
            path(
                "export-all-csv/",
                self.admin_site.admin_view(self.export_all_csv),
                name="audits_websiteaudit_export_all_csv",
            ),
            path(
                "export-all-excel/",
                self.admin_site.admin_view(self.export_all_excel),
                name="audits_websiteaudit_export_all_excel",
            ),
        ]
        return custom + urls

    def export_csv(self, request, object_id):
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("WebsiteAudit not found")
        if not (self.has_view_permission(request, obj) or self.has_change_permission(request, obj)):
            raise PermissionDenied

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (  # pyright: ignore
            f"attachment; filename=websiteaudit_{obj.pk}.csv"
        )

        writer = csv.writer(response)
        fields = [
            "id",
            "url",
            "fetched_at",
            "status_code",
            "elapsed_ms",
            "page_title",
            "score",
            "results",
            "raw",
            "rendered",
            "ia",
        ]
        writer.writerow(fields)
        writer.writerow(
            [
                obj.id,
                obj.url,
                _format_dt(obj.fetched_at),
                obj.status_code,
                obj.elapsed_ms,
                obj.page_title,
                obj.score,
                _jsonify(obj.results),
                obj.raw,
                obj.rendered,
                obj.ia,
            ]
        )
        return response

    def change_view(self, request, object_id, form_url="", extra_context=None):
        ctx: Dict[str, Any] = extra_context or {}
        ctx["export_url"] = reverse(
            "admin:audits_websiteaudit_export_csv", args=[object_id]
        )
        return super().change_view(
            request, object_id, form_url=form_url, extra_context=ctx
        )

    def changelist_view(self, request, extra_context=None):
        ctx: Dict[str, Any] = extra_context or {}
        ctx["export_all_csv_url"] = reverse("admin:audits_websiteaudit_export_all_csv")
        ctx["export_all_excel_url"] = reverse("admin:audits_websiteaudit_export_all_excel")
        return super().changelist_view(request, extra_context=ctx)

    def export_selected_csv(self, request, queryset):
        """Acción para exportar registros seleccionados a CSV"""
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=website_audits.csv"  # pyright: ignore

        writer = csv.writer(response)
        fields = ["id", "url", "fetched_at", "status_code", "elapsed_ms", 
                  "page_title", "score", "results", "raw", "rendered", "ia"]
        writer.writerow(fields)

        for obj in queryset:
            writer.writerow([
                obj.id, obj.url, _format_dt(obj.fetched_at), obj.status_code,
                obj.elapsed_ms, obj.page_title, obj.score, _jsonify(obj.results),
                obj.raw, obj.rendered, obj.ia
            ])

        self.message_user(request, f"{queryset.count()} auditorías exportadas a CSV.")
        return response

    export_selected_csv.short_description = "Exportar seleccionadas a CSV"  # pyright: ignore

    def export_selected_excel(self, request, queryset):
        """Acción para exportar registros seleccionados a Excel"""
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.title = "Website Audits"

        # Estilo de encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Encabezados
        headers = ["ID", "URL", "Fecha", "Status", "Tiempo (ms)", 
                   "Título", "Score", "Resultados", "Raw", "Rendered", "IA"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Datos
        for row_idx, obj in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=obj.id)  # pyright: ignore
            ws.cell(row=row_idx, column=2, value=obj.url)
            ws.cell(row=row_idx, column=3, value=_format_dt(obj.fetched_at))
            ws.cell(row=row_idx, column=4, value=obj.status_code)
            ws.cell(row=row_idx, column=5, value=obj.elapsed_ms)
            ws.cell(row=row_idx, column=6, value=obj.page_title)
            ws.cell(row=row_idx, column=7, value=obj.score)
            ws.cell(row=row_idx, column=8, value=_jsonify(obj.results))
            ws.cell(row=row_idx, column=9, value=obj.raw)
            ws.cell(row=row_idx, column=10, value=obj.rendered)
            ws.cell(row=row_idx, column=11, value=obj.ia)

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 10

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=website_audits.xlsx"  # pyright: ignore
        wb.save(response)

        self.message_user(request, f"{queryset.count()} auditorías exportadas a Excel.")
        return response

    export_selected_excel.short_description = "Exportar seleccionadas a Excel"  # pyright: ignore

    def export_all_csv(self, request):
        """Exportar TODAS las auditorías a CSV"""
        if not self.has_view_permission(request):
            raise PermissionDenied

        queryset = WebsiteAudit.objects.all()
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=all_website_audits.csv"  # pyright: ignore

        writer = csv.writer(response)
        fields = ["id", "url", "fetched_at", "status_code", "elapsed_ms", 
                  "page_title", "score", "results", "raw", "rendered", "ia"]
        writer.writerow(fields)

        for obj in queryset:
            writer.writerow([
                obj.id,  # pyright: ignore
                obj.url, _format_dt(obj.fetched_at), obj.status_code,
                obj.elapsed_ms, obj.page_title, obj.score, _jsonify(obj.results),
                obj.raw, obj.rendered, obj.ia
            ])

        return response

    def export_all_excel(self, request):
        """Exportar TODAS las auditorías y sus resultados a Excel"""
        if not self.has_view_permission(request):
            raise PermissionDenied

        wb = Workbook()
        
        # Hoja 1: Auditorías
        ws_audits = cast(Worksheet, wb.active)
        ws_audits.title = "Auditorías"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Encabezados auditorías
        headers_audit = ["ID", "URL", "Fecha", "Status", "Tiempo (ms)", 
                        "Título", "Score", "Raw", "Rendered", "IA"]
        for col, header in enumerate(headers_audit, 1):
            cell = ws_audits.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Datos auditorías
        audits = WebsiteAudit.objects.all()
        for row_idx, obj in enumerate(audits, 2):
            ws_audits.cell(row=row_idx, column=1, value=obj.id)  # pyright: ignore
            ws_audits.cell(row=row_idx, column=2, value=obj.url)
            ws_audits.cell(row=row_idx, column=3, value=_format_dt(obj.fetched_at))
            ws_audits.cell(row=row_idx, column=4, value=obj.status_code)
            ws_audits.cell(row=row_idx, column=5, value=obj.elapsed_ms)
            ws_audits.cell(row=row_idx, column=6, value=obj.page_title)
            ws_audits.cell(row=row_idx, column=7, value=obj.score)
            ws_audits.cell(row=row_idx, column=8, value=obj.raw)
            ws_audits.cell(row=row_idx, column=9, value=obj.rendered)
            ws_audits.cell(row=row_idx, column=10, value=obj.ia)

        # Ajustar anchos
        ws_audits.column_dimensions['A'].width = 8
        ws_audits.column_dimensions['B'].width = 50
        ws_audits.column_dimensions['C'].width = 22
        ws_audits.column_dimensions['F'].width = 40

        # Hoja 2: Resultados
        ws_results = cast(Worksheet, wb.create_sheet(title="Resultados"))
        
        headers_result = ["ID", "Auditoría ID", "Código", "Título", "Nivel", 
                         "Principio", "Veredicto", "Fuente", "Score", "Score Hint"]
        for col, header in enumerate(headers_result, 1):
            cell = ws_results.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Datos resultados
        results = WebsiteAuditResult.objects.all()
        for row_idx, obj in enumerate(results, 2):
            ws_results.cell(row=row_idx, column=1, value=obj.id)  # pyright: ignore
            ws_results.cell(row=row_idx, column=2, value=obj.audit_id)  # pyright: ignore
            ws_results.cell(row=row_idx, column=3, value=obj.code)
            ws_results.cell(row=row_idx, column=4, value=obj.title)
            ws_results.cell(row=row_idx, column=5, value=obj.level)
            ws_results.cell(row=row_idx, column=6, value=obj.principle)
            ws_results.cell(row=row_idx, column=7, value=obj.verdict)
            ws_results.cell(row=row_idx, column=8, value=obj.source)
            ws_results.cell(row=row_idx, column=9, value=obj.score)
            ws_results.cell(row=row_idx, column=10, value=obj.score_hint)

        # Ajustar anchos
        ws_results.column_dimensions['C'].width = 12
        ws_results.column_dimensions['D'].width = 30
        ws_results.column_dimensions['F'].width = 20

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=fisichecker_completo.xlsx"  # pyright: ignore
        wb.save(response)

        return response


@admin.register(WebsiteAuditResult)
class WebsiteAuditResultAdmin(admin.ModelAdmin):
    list_display = ("audit", "code", "level", "principle", "verdict", "source")
    list_filter = ("level", "principle", "verdict", "source")
    search_fields = ("code", "title", "audit__url")
    actions = ["export_selected_csv", "export_selected_excel"]

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "<path:object_id>/export-csv/",
                self.admin_site.admin_view(self.export_csv),
                name="audits_websiteauditresult_export_csv",
            ),
            path(
                "export-all-csv/",
                self.admin_site.admin_view(self.export_all_csv),
                name="audits_websiteauditresult_export_all_csv",
            ),
            path(
                "export-all-excel/",
                self.admin_site.admin_view(self.export_all_excel),
                name="audits_websiteauditresult_export_all_excel",
            ),
        ]
        return custom + urls

    def export_csv(self, request, object_id):
        obj = self.get_object(request, object_id)
        if obj is None:
            raise Http404("WebsiteAuditResult not found")
        if not (self.has_view_permission(request, obj) or self.has_change_permission(request, obj)):
            raise PermissionDenied

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (  # pyright: ignore
            f"attachment; filename=websiteauditresult_{obj.pk}.csv"
        )

        writer = csv.writer(response)
        fields = [
            "id",
            "audit_id",
            "code",
            "title",
            "level",
            "principle",
            "verdict",
            "source",
            "score",
            "score_hint",
            "details",
        ]
        writer.writerow(fields)
        writer.writerow(
            [
                obj.id,
                obj.audit_id,
                obj.code,
                obj.title,
                obj.level,
                obj.principle,
                obj.verdict,
                obj.source,
                obj.score,
                obj.score_hint,
                _jsonify(obj.details),
            ]
        )
        return response

    def change_view(self, request, object_id, form_url="", extra_context=None):
        ctx: Dict[str, Any] = extra_context or {}
        ctx["export_url"] = reverse(
            "admin:audits_websiteauditresult_export_csv", args=[object_id]
        )
        return super().change_view(
            request, object_id, form_url=form_url, extra_context=ctx
        )

    def changelist_view(self, request, extra_context=None):
        ctx: Dict[str, Any] = extra_context or {}
        ctx["export_all_csv_url"] = reverse("admin:audits_websiteauditresult_export_all_csv")
        ctx["export_all_excel_url"] = reverse("admin:audits_websiteauditresult_export_all_excel")
        return super().changelist_view(request, extra_context=ctx)

    def export_selected_csv(self, request, queryset):
        """Acción para exportar registros seleccionados a CSV"""
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=audit_results.csv"  # pyright: ignore

        writer = csv.writer(response)
        fields = ["id", "audit_id", "code", "title", "level", "principle", 
                  "verdict", "source", "score", "score_hint", "details"]
        writer.writerow(fields)

        for obj in queryset:
            writer.writerow([
                obj.id,  # pyright: ignore
                obj.audit_id,  # pyright: ignore
                obj.code, obj.title, obj.level,
                obj.principle, obj.verdict, obj.source, obj.score,
                obj.score_hint, _jsonify(obj.details)
            ])

        self.message_user(request, f"{queryset.count()} resultados exportados a CSV.")
        return response

    export_selected_csv.short_description = "Exportar seleccionados a CSV"  # pyright: ignore

    def export_selected_excel(self, request, queryset):
        """Acción para exportar registros seleccionados a Excel"""
        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.title = "Audit Results"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["ID", "Auditoría ID", "Código", "Título", "Nivel", 
                   "Principio", "Veredicto", "Fuente", "Score", "Score Hint"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, obj in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=obj.id)  # pyright: ignore
            ws.cell(row=row_idx, column=2, value=obj.audit_id)  # pyright: ignore
            ws.cell(row=row_idx, column=3, value=obj.code)
            ws.cell(row=row_idx, column=4, value=obj.title)
            ws.cell(row=row_idx, column=5, value=obj.level)
            ws.cell(row=row_idx, column=6, value=obj.principle)
            ws.cell(row=row_idx, column=7, value=obj.verdict)
            ws.cell(row=row_idx, column=8, value=obj.source)
            ws.cell(row=row_idx, column=9, value=obj.score)
            ws.cell(row=row_idx, column=10, value=obj.score_hint)

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['F'].width = 20

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=audit_results.xlsx"  # pyright: ignore
        wb.save(response)

        self.message_user(request, f"{queryset.count()} resultados exportados a Excel.")
        return response

    export_selected_excel.short_description = "Exportar seleccionados a Excel"  # pyright: ignore

    def export_all_csv(self, request):
        """Exportar TODOS los resultados a CSV"""
        if not self.has_view_permission(request):
            raise PermissionDenied

        queryset = WebsiteAuditResult.objects.all()
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=all_audit_results.csv"  # pyright: ignore

        writer = csv.writer(response)
        fields = ["id", "audit_id", "code", "title", "level", "principle", 
                  "verdict", "source", "score", "score_hint", "details"]
        writer.writerow(fields)

        for obj in queryset:
            writer.writerow([
                obj.id,  # pyright: ignore
                obj.audit_id,  # pyright: ignore
                obj.code, obj.title, obj.level,
                obj.principle, obj.verdict, obj.source, obj.score,
                obj.score_hint, _jsonify(obj.details)
            ])

        return response

    def export_all_excel(self, request):
        """Exportar TODOS los resultados a Excel"""
        if not self.has_view_permission(request):
            raise PermissionDenied

        wb = Workbook()
        ws = cast(Worksheet, wb.active)
        ws.title = "Resultados"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["ID", "Auditoría ID", "Código", "Título", "Nivel", 
                   "Principio", "Veredicto", "Fuente", "Score", "Score Hint"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        results = WebsiteAuditResult.objects.all()
        for row_idx, obj in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=obj.id)  # pyright: ignore
            ws.cell(row=row_idx, column=2, value=obj.audit_id)  # pyright: ignore
            ws.cell(row=row_idx, column=3, value=obj.code)
            ws.cell(row=row_idx, column=4, value=obj.title)
            ws.cell(row=row_idx, column=5, value=obj.level)
            ws.cell(row=row_idx, column=6, value=obj.principle)
            ws.cell(row=row_idx, column=7, value=obj.verdict)
            ws.cell(row=row_idx, column=8, value=obj.source)
            ws.cell(row=row_idx, column=9, value=obj.score)
            ws.cell(row=row_idx, column=10, value=obj.score_hint)

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['F'].width = 20

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=all_audit_results.xlsx"  # pyright: ignore
        wb.save(response)

        return response
