# audits/views.py
from typing import Any, Mapping, Optional, cast
import traceback
import csv
import json

from django.conf import settings
from django.db import connection
from django.db.utils import OperationalError
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.generics import ListAPIView, RetrieveAPIView
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.decorators import action
from rest_framework import exceptions
from rest_framework.authentication import SessionAuthentication

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .serializers import WebsiteAuditListSerializer, WebsiteAuditSerializer, AuditRequestSerializer
from .models import WebsiteAudit
from .audit import scrape_and_audit
from .utils.persist import persist_audit_with_results
from .checks.criteria.base import CheckMode
from .statistics import AuditStatistics

def _parse_mode(v: str) -> CheckMode:
    v2 = (v or "").strip().lower()
    if v2 in ("rendered", "r"):
        return CheckMode.RENDERED
    if v2 in ("ai",):
        return CheckMode.AI
    if v2 in ("auto", "a"):
        return CheckMode.AUTO
    return CheckMode.RAW

def _fmt_detail_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)

def _get_details_dict(details):
    if isinstance(details, dict):
        return details
    if isinstance(details, str):
        try:
            return json.loads(details)
        except Exception:
            return {}
    return {}

def _results_accessor_name() -> Optional[str]:
    # Busca el accessor reverse hacia WebsiteAuditResult (p.ej. "websiteauditresult_set")
    for rel in WebsiteAudit._meta.related_objects:
        if rel.related_model is AuditRequestSerializer:
            return rel.get_accessor_name()
    return None

@csrf_exempt
class AuditView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = [SessionAuthentication]

    def post(self, request, *args, **kwargs):
        s = AuditRequestSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        vd: Mapping[str, Any] = cast(Mapping[str, Any], getattr(s, "validated_data", {}))
        url = vd.get("url")

        if not isinstance(url, str) or not url.strip():
            return Response({"detail": "Campo 'url' requerido o inválido."}, status=status.HTTP_400_BAD_REQUEST)

        # Modo y flag IA (query o body)
        mode_str = str(request.query_params.get("mode", request.data.get("mode", "raw")))
        run_mode = _parse_mode(mode_str)
        ai_flag = str(request.query_params.get("ai", request.data.get("ai", ""))).lower() in ("1", "true", "yes")

        # Cierra DB antes del trabajo largo
        try: connection.close()
        except Exception: pass

        try:
            data = scrape_and_audit(url, mode=run_mode, use_ai=ai_flag)
        except Exception as e:
            payload = {"detail": str(e)}
            if settings.DEBUG:
                payload["where"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_400_BAD_REQUEST)

        wcag = data.get("wcag") or data.get("results") or {}
        try:
            # Determinar usuario (solo si está autenticado)
            user = request.user if request.user.is_authenticated else None
            
            audit = persist_audit_with_results(
                url=url,
                response_meta=data,
                wcag=wcag,
                rendered=bool(data.get("rendered", False)),
                rendered_codes=data.get("rendered_codes"),
                user=user,
            )
        except OperationalError as e:
            return Response(
                {"detail": "Database unavailable while saving audit.", "db_error": str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )
        except Exception as e:
            payload = {"detail": str(e)}
            if settings.DEBUG:
                payload["where"] = traceback.format_exc()
            return Response(payload, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(WebsiteAuditSerializer(audit).data, status=status.HTTP_201_CREATED)


class AuditRetrieveView(RetrieveAPIView):
    queryset = WebsiteAudit.objects.all()
    serializer_class = WebsiteAuditSerializer


class AuditListView(ListAPIView):
    permission_classes = [AllowAny]
    authentication_classes = [SessionAuthentication]
    serializer_class = WebsiteAuditListSerializer

    def get_queryset(self):
        # Debug: Imprimir información de autenticación
        print(f"[DEBUG] Usuario autenticado: {self.request.user}")
        print(f"[DEBUG] Is authenticated: {self.request.user.is_authenticated}")
        
        # Si el usuario está autenticado, solo mostrar sus auditorías
        if self.request.user.is_authenticated:
            qs = WebsiteAudit.objects.filter(user=self.request.user).order_by("-fetched_at")
            print(f"[DEBUG] Auditorías del usuario {self.request.user.username}: {qs.count()}")
        else:
            # Si no está autenticado, mostrar auditorías sin usuario asignado
            qs = WebsiteAudit.objects.filter(user__isnull=True).order_by("-fetched_at")
            print(f"[DEBUG] Auditorías sin usuario: {qs.count()}")
        acc = _results_accessor_name()
        if acc:
            try:
                qs = qs.prefetch_related(acc)
            except Exception:
                # Si por algún motivo falla, seguimos sin prefetch para no romper la vista
                pass
        return qs


# ============= VISTAS DE ESTADÍSTICAS =============

class StatisticsGlobalView(APIView):
    """
    GET /api/statistics/global/
    Retorna estadísticas globales de todas las auditorías.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        stats = AuditStatistics.get_global_statistics()
        return Response(stats)


class StatisticsVerdictDistributionView(APIView):
    """
    GET /api/statistics/verdicts/
    Retorna la distribución de veredictos (pass, fail, partial, na).
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        stats = AuditStatistics.get_verdict_distribution()
        return Response(stats)


class StatisticsCriteriaView(APIView):
    """
    GET /api/statistics/criteria/
    Retorna estadísticas por criterio WCAG.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        stats = AuditStatistics.get_criteria_statistics()
        return Response(stats)


class StatisticsLevelView(APIView):
    """
    GET /api/statistics/levels/
    Retorna estadísticas agrupadas por nivel (A, AA, AAA).
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        stats = AuditStatistics.get_level_statistics()
        return Response(stats)


class StatisticsPrincipleView(APIView):
    """
    GET /api/statistics/principles/
    Retorna estadísticas agrupadas por principio WCAG.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        stats = AuditStatistics.get_principle_statistics()
        return Response(stats)


class StatisticsTimelineView(APIView):
    """
    GET /api/statistics/timeline/?days=30
    Retorna estadísticas de auditorías a lo largo del tiempo.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        days = int(request.query_params.get('days', 30))
        stats = AuditStatistics.get_timeline_statistics(days)
        return Response(stats)


class StatisticsURLRankingView(APIView):
    """
    GET /api/statistics/ranking/?limit=10
    Retorna el ranking de URLs mejor y peor puntuadas.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        limit = int(request.query_params.get('limit', 10))
        stats = AuditStatistics.get_url_ranking(limit)
        return Response(stats)


class StatisticsSourceComparisonView(APIView):
    """
    GET /api/statistics/sources/
    Compara resultados entre diferentes fuentes (raw, rendered, mixed).
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        stats = AuditStatistics.get_source_comparison()
        return Response(stats)


class StatisticsAuditDetailView(APIView):
    """
    GET /api/statistics/audit/<id>/
    Retorna estadísticas detalladas para una auditoría específica.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request, audit_id):
        stats = AuditStatistics.get_detailed_audit_statistics(audit_id)
        if "error" in stats:
            return Response(stats, status=status.HTTP_404_NOT_FOUND)
        return Response(stats)


class StatisticsComprehensiveReportView(APIView):
    """
    GET /api/statistics/report/
    Genera un reporte completo con todas las estadísticas principales.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        report = AuditStatistics.get_comprehensive_report()
        return Response(report)


class StatisticsAccessibilityLevelsView(APIView):
    """
    GET /api/statistics/accessibility-levels/
    Calcula niveles de accesibilidad según Hilera et al. (2013).
    Clasifica sitios en: Alto, Moderado, Deficiente, Muy deficiente.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        stats = AuditStatistics.get_accessibility_levels_hilera()
        return Response(stats)


class StatisticsAccessibilityByWCAGView(APIView):
    """
    GET /api/statistics/accessibility-by-wcag/
    Calcula niveles de accesibilidad por nivel WCAG (A, AA, AAA).
    Incluye el promedio de cumplimiento de los 3 niveles.
    """
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        stats = AuditStatistics.get_accessibility_levels_by_wcag_level()
        return Response(stats)


# ============================================
# Vistas de Autenticación
# ============================================


@api_view(['POST'])
@permission_classes([AllowAny])
@csrf_exempt
def login_view(request):
    """
    POST /api/auth/login
    Body: { "username": "...", "password": "..." }
    Retorna: { "token": "session", "user": { "id", "username", "email" } }
    """
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
        return Response(
            {'detail': 'Usuario y contraseña requeridos'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    user = authenticate(request, username=username, password=password)
    
    if user is not None:
        login(request, user)
        # Type assertion para Pylance - authenticate retorna User o None
        assert isinstance(user, User)
        return Response({
            'token': 'session',  # Usando sesiones de Django
            'user': {
                'id': user.pk,
                'username': user.username,
                'email': user.email,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser,
            }
        }, status=status.HTTP_200_OK)
    
    return Response(
        {'detail': 'Credenciales inválidas'}, 
        status=status.HTTP_401_UNAUTHORIZED
    )


@csrf_exempt
@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])  # Permitir logout sin autenticación estricta
def logout_view(request):
    """
    POST /api/auth/logout
    Cierra la sesión del usuario actual y elimina las cookies.
    """
    from django.conf import settings
    
    # Ejecutar logout de Django
    logout(request)
    
    # Crear respuesta
    response = Response({
        'detail': 'Sesión cerrada exitosamente',
        'message': '¡Hasta pronto! Tu sesión ha sido cerrada de forma segura.'
    }, status=status.HTTP_200_OK)
    
    # Eliminar cookies con los mismos parámetros que se usaron para crearlas
    response.delete_cookie(
        settings.SESSION_COOKIE_NAME,
        path=settings.SESSION_COOKIE_PATH,
        domain=settings.SESSION_COOKIE_DOMAIN,
        samesite=settings.SESSION_COOKIE_SAMESITE,
    )
    response.delete_cookie(
        settings.CSRF_COOKIE_NAME,
        path=settings.CSRF_COOKIE_PATH,
        domain=settings.SESSION_COOKIE_DOMAIN,
        samesite=settings.CSRF_COOKIE_SAMESITE,
    )
    
    return response


@api_view(['GET'])
@permission_classes([AllowAny])
@ensure_csrf_cookie
def current_user_view(request):
    """
    GET /api/auth/user
    Retorna información del usuario autenticado actual.
    Si no está autenticado, devuelve 401.
    """
    if not request.user.is_authenticated:
        return Response(
            {'detail': 'No autenticado'}, 
            status=status.HTTP_401_UNAUTHORIZED
        )
    
    user: User = request.user
    return Response({
        'id': user.pk,
        'username': user.username,
        'email': user.email,
        'is_staff': user.is_staff,
        'is_superuser': user.is_superuser,
    }, status=status.HTTP_200_OK)

# ============================================
# Vista para borrar auditoría y sus resultados
# ============================================

@csrf_exempt
@api_view(['DELETE'])
@authentication_classes([SessionAuthentication])
@permission_classes([IsAuthenticated])
def delete_audit_view(request, audit_id):
    """
    DELETE /api/audits/<audit_id>/delete/
    Borra la auditoría y todos sus resultados asociados.
    """
    user: User = request.user
    try:
        audit = WebsiteAudit.objects.get(pk=audit_id, user=user)
    except WebsiteAudit.DoesNotExist:
        return Response({'detail': 'Auditoría no encontrada o no pertenece al usuario.'}, status=status.HTTP_404_NOT_FOUND)

    # Borrar todos los resultados asociados primero (por seguridad, aunque CASCADE)
    audit.criterion_results.all().delete() # type: ignore[attr-defined]
    audit.delete()
    return Response({'detail': 'Auditoría y resultados borrados correctamente.'}, status=status.HTTP_200_OK)

# ============================================
# Vistas de Exportación
# ============================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_csv_view(request):
    """
    GET /api/export/csv
    Exporta las auditorías del usuario autenticado a CSV con detalles de criterios.
    """
    user: User = request.user
    audits = WebsiteAudit.objects.filter(user=user).prefetch_related('criterion_results').order_by('-fetched_at')
    
    # Crear respuesta HTTP con CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="auditorias_detalladas_{user.username}.csv"'
    
    # Escribir CSV con BOM para Excel
    response.write('\ufeff')
    writer = csv.writer(response)
    
    # SECCIÓN 1: Resumen de Auditorías
    writer.writerow(['=== RESUMEN DE AUDITORÍAS ==='])
    writer.writerow([
        'ID', 'URL', 'Título', 'Score (%)', 'HTTP Status', 
        'Fecha', 'Tiempo (ms)', 'Modo', 'Total Criterios', 'Pass', 'Fail', 'Partial', 'N/A'
    ])
    
    for audit in audits:
        mode = 'RAW'
        if audit.ia:
            mode = 'AI'
        elif audit.rendered:
            mode = 'RENDERED'
        
        score_pct = round(audit.score * 100, 1) if audit.score is not None else ''
        
        # Contar veredictos
        results = audit.criterion_results.all()  # type: ignore[attr-defined]
        total = results.count()
        pass_count = results.filter(verdict='pass').count()
        fail_count = results.filter(verdict='fail').count()
        partial_count = results.filter(verdict='partial').count()
        na_count = results.filter(verdict='na').count()
        
        writer.writerow([
            audit.id,
            audit.url,
            audit.page_title or '',
            score_pct,
            audit.status_code or '',
            audit.fetched_at.strftime('%Y-%m-%d %H:%M:%S'),
            audit.elapsed_ms or '',
            mode,
            total,
            pass_count,
            fail_count,
            partial_count,
            na_count,
        ])
    
    # SECCIÓN 2: Detalles de Criterios por Auditoría
    writer.writerow([])
    writer.writerow(['=== DETALLES DE CRITERIOS WCAG ==='])
    writer.writerow([
        'Auditoría ID', 'URL', 'Criterio', 'Título', 'Nivel', 'Principio', 
        'Veredicto', 'Score', 'Evidencias', 'Errores', 'Advertencias'
    ])
    
    for audit in audits:
        for result in audit.criterion_results.all(): # type: ignore[attr-defined]
            details = _get_details_dict(result.details)
            writer.writerow([
                audit.id,
                audit.url,
                result.code,
                result.title,
                result.level,
                result.principle,
                result.verdict,
                result.score if result.score is not None else '',
                _fmt_detail_value(details.get('evidences')),
                _fmt_detail_value(details.get('errors')),
                _fmt_detail_value(details.get('warnings')),
            ])
    
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel_view(request):
    """
    GET /api/export/excel
    Exporta las auditorías del usuario autenticado a Excel con múltiples hojas.
    """
    user: User = request.user
    audits = WebsiteAudit.objects.filter(user=user).prefetch_related('criterion_results').order_by('-fetched_at')
    
    # Crear workbook
    wb = Workbook()
    
    # ===== HOJA 1: RESUMEN DE AUDITORÍAS =====
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados Resumen
    headers_summary = [
        'ID', 'URL', 'Título', 'Score (%)', 'HTTP Status', 
        'Fecha', 'Tiempo (ms)', 'Modo', 'Total', 'Pass', 'Fail', 'Partial', 'N/A'
    ]
    
    for col, header in enumerate(headers_summary, start=1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Datos Resumen
    for row_idx, audit in enumerate(audits, start=2):
        mode = 'RAW'
        if audit.ia:
            mode = 'AI'
        elif audit.rendered:
            mode = 'RENDERED'
        
        score_pct = round(audit.score * 100, 1) if audit.score is not None else ''
        
        # Contar veredictos
        results = audit.criterion_results.all()  # type: ignore[attr-defined]
        total = results.count()
        pass_count = results.filter(verdict='pass').count()
        fail_count = results.filter(verdict='fail').count()
        partial_count = results.filter(verdict='partial').count()
        na_count = results.filter(verdict='na').count()
        
        ws_summary.cell(row=row_idx, column=1, value=audit.id)
        ws_summary.cell(row=row_idx, column=2, value=audit.url)
        ws_summary.cell(row=row_idx, column=3, value=audit.page_title or '')
        ws_summary.cell(row=row_idx, column=4, value=score_pct)
        ws_summary.cell(row=row_idx, column=5, value=audit.status_code or '')
        ws_summary.cell(row=row_idx, column=6, value=audit.fetched_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws_summary.cell(row=row_idx, column=7, value=audit.elapsed_ms or '')
        ws_summary.cell(row=row_idx, column=8, value=mode)
        ws_summary.cell(row=row_idx, column=9, value=total)
        ws_summary.cell(row=row_idx, column=10, value=pass_count)
        ws_summary.cell(row=row_idx, column=11, value=fail_count)
        ws_summary.cell(row=row_idx, column=12, value=partial_count)
        ws_summary.cell(row=row_idx, column=13, value=na_count)
    
    # Ajustar anchos Resumen
    ws_summary.column_dimensions['A'].width = 8
    ws_summary.column_dimensions['B'].width = 50
    ws_summary.column_dimensions['C'].width = 40
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 20
    ws_summary.column_dimensions['G'].width = 15
    ws_summary.column_dimensions['H'].width = 12
    ws_summary.column_dimensions['I'].width = 10
    ws_summary.column_dimensions['J'].width = 10
    ws_summary.column_dimensions['K'].width = 10
    ws_summary.column_dimensions['L'].width = 10
    ws_summary.column_dimensions['M'].width = 10
    
    # ===== HOJA 2: DETALLES DE CRITERIOS =====
    ws_details = wb.create_sheet(title="Criterios WCAG")
    
    # Encabezados Detalles
    headers_details = [
        'Auditoría ID', 'URL', 'Criterio', 'Título', 'Nivel', 'Principio',
        'Veredicto', 'Score', 'Evidencias', 'Errores', 'Advertencias'
    ]
    
    for col, header in enumerate(headers_details, start=1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    # Datos Detalles
    detail_row = 2
    for audit in audits:
        for result in audit.criterion_results.all():
            details = _get_details_dict(result.details)
            ws_details.cell(row=detail_row, column=1, value=audit.id)
            ws_details.cell(row=detail_row, column=2, value=audit.url)
            ws_details.cell(row=detail_row, column=3, value=result.code)
            ws_details.cell(row=detail_row, column=4, value=result.title)
            ws_details.cell(row=detail_row, column=5, value=result.level)
            ws_details.cell(row=detail_row, column=6, value=result.principle)
            ws_details.cell(row=detail_row, column=7, value=result.verdict)
            ws_details.cell(row=detail_row, column=8, value=result.score if result.score is not None else '')
            ws_details.cell(row=detail_row, column=9, value=_fmt_detail_value(details.get('evidences')))
            ws_details.cell(row=detail_row, column=10, value=_fmt_detail_value(details.get('errors')))
            ws_details.cell(row=detail_row, column=11, value=_fmt_detail_value(details.get('warnings')))
            detail_row += 1
    
    # Ajustar anchos Detalles
    ws_details.column_dimensions['A'].width = 12
    ws_details.column_dimensions['B'].width = 50
    ws_details.column_dimensions['C'].width = 10
    ws_details.column_dimensions['D'].width = 50
    ws_details.column_dimensions['E'].width = 8
    ws_details.column_dimensions['F'].width = 18
    ws_details.column_dimensions['G'].width = 12
    ws_details.column_dimensions['H'].width = 10
    ws_details.column_dimensions['I'].width = 15
    ws_details.column_dimensions['J'].width = 15
    ws_details.column_dimensions['K'].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="auditorias_detalladas_{user.username}.xlsx"'
    
    # Guardar workbook en response
    wb.save(response)
    
    return response
